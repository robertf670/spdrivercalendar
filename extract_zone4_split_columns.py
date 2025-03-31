import pandas as pd
import os
import re
from openpyxl import load_workbook
import logging

# Set up logging
logging.basicConfig(level=logging.INFO, 
                    format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def convert_zone4_to_csv(excel_file, output_dir=None):
    """
    Convert Zone4 Excel sheets to CSV files matching the Zone3 format.
    Handles the dual-column layout where duties are arranged side by side.
    """
    # Get base filename
    base_name = os.path.splitext(os.path.basename(excel_file))[0]
    
    # Set output directory
    if output_dir is None:
        output_dir = os.path.dirname(excel_file)
    
    # Ensure output directory exists
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    logger.info(f"Reading Excel file: {excel_file}")
    
    # Load workbook using openpyxl with data_only=True to get values instead of formulas
    wb = load_workbook(filename=excel_file, data_only=True)
    
    logger.info(f"Sheets in workbook: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames:
        # Skip non-schedule sheets
        if "roster" in sheet_name.lower() or "summary" in sheet_name.lower():
            logger.info(f"Skipping sheet '{sheet_name}' - not a schedule sheet")
            continue
        
        # Determine CSV suffix based on sheet name
        if "M-F" in sheet_name.upper() or "M-f" in sheet_name.upper() or "MONDAY" in sheet_name.upper():
            csv_suffix = "MF"
        elif "SAT" in sheet_name.upper() or "SATURDAY" in sheet_name.upper():
            csv_suffix = "Sat" 
        elif "SUN" in sheet_name.upper() or "SUNDAY" in sheet_name.upper():
            csv_suffix = "Sun"
        else:
            csv_suffix = sheet_name.replace(" ", "_")
        
        logger.info(f"Processing sheet '{sheet_name}' with suffix {csv_suffix}...")
        
        # Process the sheet to extract duty data, handling the split columns
        duty_data = process_sheet_dual_columns(wb[sheet_name])
        
        if duty_data:
            logger.info(f"Found {len(duty_data)} duties in sheet '{sheet_name}'")
            
            # Convert to CSV format
            csv_data = convert_to_csv_format(duty_data)
            
            # Create and save the CSV file
            csv_filename = f"{base_name}{csv_suffix}.csv"
            csv_path = os.path.join(output_dir, csv_filename)
            
            # Create DataFrame and save to CSV
            df = pd.DataFrame(csv_data)
            df.to_csv(csv_path, index=False)
            logger.info(f"Saved sheet '{sheet_name}' to {csv_path}")
        else:
            logger.warning(f"No duty data extracted from sheet '{sheet_name}'")
    
    logger.info("Conversion complete!")

def process_sheet_dual_columns(worksheet):
    """
    Process a worksheet with a dual-column layout to extract duty information.
    The Excel has a structure like: left_route, left_place, left_arr, left_dep, right_route, right_place, right_arr, right_dep
    """
    logger.info(f"Processing dual-column worksheet: {worksheet.title}")
    
    # First, identify the column structure by finding the headers (Route, Place, Arr, Dep)
    column_indices = identify_column_structure(worksheet)
    if not column_indices:
        logger.warning(f"Could not identify column structure in sheet {worksheet.title}")
        return []
    
    logger.info(f"Identified column structure: {column_indices}")
    
    # Find all duties in the sheet
    duty_sections = find_duty_sections(worksheet)
    logger.info(f"Found {len(duty_sections)} duty sections")
    
    # Extract data from each duty section
    duties = []
    for section in duty_sections:
        duty_number = section['duty_number']
        reports_time = section['reports']
        
        # Process left side
        left_duty_entries = process_duty_side(
            worksheet, 
            section, 
            column_indices['left_route'], 
            column_indices['left_place'], 
            column_indices['left_arr'], 
            column_indices['left_dep'],
            'left',
            duty_number
        )
        
        if left_duty_entries:
            duties.append({
                'duty_number': duty_number,
                'reports': reports_time,
                'entries': left_duty_entries
            })
        
        # Process right side - if there's a right-side duty in this section
        right_duty = find_right_duty(worksheet, section, column_indices['right_route'])
        if right_duty:
            right_duty_number = right_duty['duty_number']
            right_reports_time = right_duty['reports']
            
            right_duty_entries = process_duty_side(
                worksheet, 
                section, 
                column_indices['right_route'], 
                column_indices['right_place'], 
                column_indices['right_arr'], 
                column_indices['right_dep'],
                'right',
                right_duty_number
            )
            
            if right_duty_entries:
                duties.append({
                    'duty_number': right_duty_number,
                    'reports': right_reports_time,
                    'entries': right_duty_entries
                })
    
    # Sort duties by duty number
    duties.sort(key=lambda x: int(x['duty_number']))
    
    logger.info(f"Extracted {len(duties)} duties in total")
    return duties

def identify_column_structure(worksheet):
    """
    Identify the column indices for route, place, arr, dep on both left and right sides.
    """
    # Initialize column indices
    column_indices = {
        'left_route': None, 
        'left_place': None, 
        'left_arr': None, 
        'left_dep': None,
        'right_route': None, 
        'right_place': None, 
        'right_arr': None, 
        'right_dep': None
    }
    
    # Search for headers in the first 20 rows
    for row_idx in range(min(20, worksheet.max_row)):
        row_values = [cell.value for cell in worksheet[row_idx+1]]
        
        # Convert to strings and lowercase for easier matching
        row_text = [str(val).lower() if val else '' for val in row_values]
        
        # Look for columns by their headers
        route_indices = [i for i, val in enumerate(row_text) if 'route' in val]
        place_indices = [i for i, val in enumerate(row_text) if 'place' in val]
        arr_indices = [i for i, val in enumerate(row_text) if 'arr' in val]
        dep_indices = [i for i, val in enumerate(row_text) if 'dep' in val]
        
        # If we found all column types
        if route_indices and place_indices and arr_indices and dep_indices:
            # Must have at least 2 of each for dual columns
            if len(route_indices) >= 2 and len(place_indices) >= 2 and len(arr_indices) >= 2 and len(dep_indices) >= 2:
                # Assign left/right columns based on position
                column_indices['left_route'] = route_indices[0]
                column_indices['left_place'] = place_indices[0]
                column_indices['left_arr'] = arr_indices[0]
                column_indices['left_dep'] = dep_indices[0]
                
                column_indices['right_route'] = route_indices[1]
                column_indices['right_place'] = place_indices[1]
                column_indices['right_arr'] = arr_indices[1]
                column_indices['right_dep'] = dep_indices[1]
                
                return column_indices
    
    # Fallback to guessing columns by position if header search failed
    # Assume a typical layout of 8 columns: route,place,arr,dep,route,place,arr,dep
    if worksheet.max_column >= 8:
        column_indices['left_route'] = 0
        column_indices['left_place'] = 1
        column_indices['left_arr'] = 2
        column_indices['left_dep'] = 3
        column_indices['right_route'] = 4
        column_indices['right_place'] = 5
        column_indices['right_arr'] = 6
        column_indices['right_dep'] = 7
        
        return column_indices
    
    return None

def find_duty_sections(worksheet):
    """
    Find all duty sections in the worksheet.
    A duty section starts with a line containing "Duty XXX" and ends before the next duty.
    """
    duty_sections = []
    current_section = None
    
    for row_idx in range(1, worksheet.max_row + 1):
        row_values = [cell.value for cell in worksheet[row_idx]]
        row_text = ' '.join([str(val) if val else '' for val in row_values]).lower()
        
        # Look for duty number pattern in left columns (first half of the row)
        left_values = row_values[:len(row_values)//2] if len(row_values) > 1 else row_values
        left_text = ' '.join([str(val) if val else '' for val in left_values]).lower()
        
        duty_match = re.search(r'duty\s+(\d{3})', left_text)
        if duty_match:
            # If we're already tracking a section, save it before starting a new one
            if current_section:
                current_section['end_row'] = row_idx - 1
                duty_sections.append(current_section)
            
            duty_number = duty_match.group(1)
            
            # Look for reports time
            reports_match = re.search(r'reports at\s+(\d{2}:\d{2})', left_text)
            reports_time = reports_match.group(1) if reports_match else ''
            
            # Start a new section
            current_section = {
                'duty_number': duty_number,
                'reports': reports_time,
                'start_row': row_idx,
                'end_row': None
            }
    
    # Add the last section if we have one
    if current_section:
        current_section['end_row'] = worksheet.max_row
        duty_sections.append(current_section)
    
    return duty_sections

def find_right_duty(worksheet, section, right_route_col):
    """
    Find if there's a duty on the right side by looking for duty numbers in the right columns.
    """
    for row_idx in range(section['start_row'], section['end_row'] + 1):
        right_side_text = ''
        
        # Get text from right side columns
        for col_idx in range(right_route_col - 1, min(right_route_col + 5, worksheet.max_column + 1)):
            cell_value = worksheet.cell(row=row_idx, column=col_idx).value
            if cell_value:
                right_side_text += str(cell_value) + ' '
        
        # Look for duty number or "takes up" pattern in right side
        duty_match = re.search(r'duty\s+(\d{3})', right_side_text.lower())
        takes_up_match = re.search(r'takes up', right_side_text.lower())
        
        if duty_match:
            duty_number = duty_match.group(1)
            
            # Look for reports time
            reports_match = re.search(r'reports at\s+(\d{2}:\d{2})', right_side_text.lower())
            reports_time = reports_match.group(1) if reports_match else ''
            
            return {
                'duty_number': duty_number,
                'reports': reports_time,
                'row': row_idx
            }
        elif takes_up_match:
            # Try to extract duty number from "takes up" pattern
            duty_from_takes_up = extract_duty_from_takes_up(right_side_text)
            if duty_from_takes_up:
                return {
                    'duty_number': duty_from_takes_up,
                    'reports': '',
                    'row': row_idx
                }
    
    return None

def extract_duty_from_takes_up(text):
    """Extract duty number from a 'takes up' text"""
    # Look for "duty XXX takes up"
    duty_match = re.search(r'duty\s+(\d{3}).*?takes up', text.lower())
    if duty_match:
        return duty_match.group(1)
    
    # Look for "takes up bus XXX"
    bus_match = re.search(r'takes up.*?bus\s+(\d+)', text.lower())
    if bus_match:
        # Map bus number to duty if needed
        return None
    
    return None

def extract_side_text(worksheet, row_idx, is_right_side, column_indices):
    """
    Extract text from either the left or right side of a row
    """
    if is_right_side == 'right':
        # Get text from the right side columns
        right_cols = [
            column_indices.get('right_route', 4),
            column_indices.get('right_place', 5),
            column_indices.get('right_arr', 6), 
            column_indices.get('right_dep', 7)
        ]
        
        # Add a few columns to each side to catch any additional info
        cols_to_check = list(range(min(right_cols) - 1, max(right_cols) + 2))
        
    else:  # left side
        # Get text from the left side columns
        left_cols = [
            column_indices.get('left_route', 0),
            column_indices.get('left_place', 1),
            column_indices.get('left_arr', 2),
            column_indices.get('left_dep', 3)
        ]
        
        # Add a few columns to each side to catch any additional info
        cols_to_check = list(range(min(left_cols) - 1, max(left_cols) + 2))
    
    # Collect text from the selected columns
    side_text = ''
    for col_idx in cols_to_check:
        if col_idx >= 0 and col_idx < worksheet.max_column:
            cell_value = worksheet.cell(row=row_idx, column=col_idx+1).value
            if cell_value:
                side_text += str(cell_value) + ' '
    
    return side_text.strip().lower()

def extract_full_row_text(worksheet, row_idx):
    """Extract all text from a row to scan for special notes"""
    row_text = ''
    for cell in worksheet[row_idx]:
        if cell.value:
            row_text += str(cell.value) + ' '
    return row_text.strip().lower()

def check_special_notes(row_text, duty_number):
    """Check for special notes like 'Finish Duty' or 'Takes Bus'"""
    
    # Check for "Finish Duty" pattern
    if 'finish' in row_text and 'duty' in row_text:
        return f"Duty {duty_number} Finished Duty"
    
    # Check for "Takes Bus" pattern
    takes_bus_match = re.search(r'takes\s+bus\s+(\d+)', row_text)
    if takes_bus_match:
        bus_number = takes_bus_match.group(1)
        
        # Try to extract route information
        route_match = re.search(r'route\s+(\w+)', row_text)
        route_info = f"Route {route_match.group(1)}" if route_match else ""
        
        # Try to extract time information
        time_match = re.search(r'at\s+(\d{1,2}:\d{2})', row_text)
        time_info = f"at {time_match.group(1)}" if time_match else ""
        
        # Try to extract location
        location_match = re.search(r'at\s+\d{1,2}:\d{2}\s+(\w+)', row_text)
        location = location_match.group(1) if location_match else ""
        
        note = f"Takes Bus {bus_number}"
        if route_info:
            note += f" ({route_info})"
        if time_info:
            note += f" {time_info}"
        if location:
            note += f" {location}"
        
        return note
    
    return None

def process_duty_side(worksheet, section, route_col, place_col, arr_col, dep_col, side='left', current_duty_number=None):
    """
    Process one side (left or right) of a duty section to extract entries.
    """
    start_row = section['start_row']
    end_row = section['end_row']
    duty_entries = []
    
    if side == 'right':
        # For right side duties, find the starting row by scanning for the duty number
        for i in range(start_row, end_row + 1):
            side_text = extract_side_text(worksheet, i, 'right', {
                'right_route': route_col,
                'right_place': place_col,
                'right_arr': arr_col,
                'right_dep': dep_col
            })
            
            if f"duty {current_duty_number}" in side_text or f"takes up" in side_text:
                start_row = i
                break
    
    # Variables to track state
    next_from_location = None
    
    for row_idx in range(start_row, end_row + 1):
        # Skip header rows
        full_row_text = extract_full_row_text(worksheet, row_idx)
        if 'route' in full_row_text and 'place' in full_row_text:
            continue
        
        # Extract text from just this side (left or right) to prevent cross-contamination
        side_text = extract_side_text(worksheet, row_idx, side, {
            'left_route': 0,
            'left_place': 1,
            'left_arr': 2,
            'left_dep': 3,
            'right_route': route_col,
            'right_place': place_col,
            'right_arr': arr_col,
            'right_dep': dep_col
        })
        
        # Get values from the appropriate columns
        route_value = worksheet.cell(row=row_idx, column=route_col+1).value
        place_value = worksheet.cell(row=row_idx, column=place_col+1).value
        arr_value = worksheet.cell(row=row_idx, column=arr_col+1).value
        dep_value = worksheet.cell(row=row_idx, column=dep_col+1).value
        
        # Clean up values
        route = str(route_value).strip() if route_value else ""
        location = str(place_value).strip() if place_value else ""
        arrival = str(arr_value).strip() if arr_value else ""
        departure = str(dep_value).strip() if dep_value else ""
        
        # Check if this is a time value
        if arrival and not re.match(r'^\d{1,2}:\d{2}(:\d{2})?$', arrival):
            arrival = ""
        if departure and not re.match(r'^\d{1,2}:\d{2}(:\d{2})?$', departure):
            departure = ""
        
        # Check for "Departs garage"
        departs_garage = False
        departs_time = ""
        if "departs garage" in side_text:
            departs_garage = True
            # Find the departs time
            for cell in worksheet[row_idx]:
                if cell.value and re.match(r'^\d{1,2}:\d{2}(:\d{2})?$', str(cell.value).strip()):
                    departs_time = str(cell.value).strip()
                    break
        
        # Extract notes specific to this duty
        notes = ""
        
        # First row often has reports info for left side or "takes up" info for right side
        if row_idx == start_row:
            if side == 'left' and section['reports']:
                notes = f"Duty {current_duty_number} reports at {section['reports']}"
            elif side == 'right':
                takes_up_match = re.search(r'takes up.*?(\d{2}:\d{2})', side_text)
                if takes_up_match:
                    takes_up_time = takes_up_match.group(1)
                    location_match = re.search(r'(\d{2}:\d{2})\s+([a-z0-9]+)', side_text)
                    location_text = location_match.group(2) if location_match else ""
                    notes = f"Duty {current_duty_number} takes up at {takes_up_time}"
                    if location_text:
                        notes += f" {location_text}"
        else:
            # Check for "Takes up" notes specific to this duty
            if f"duty {current_duty_number}" in side_text:
                takes_up_match = re.search(r'takes up.*?(\d{2}:\d{2})', side_text)
                if takes_up_match:
                    takes_up_time = takes_up_match.group(1)
                    notes = f"Duty {current_duty_number} takes up at {takes_up_time}"
            
            # Check for special notes across the entire row (Finish Duty, Takes Bus, etc.)
            special_note = check_special_notes(full_row_text, current_duty_number)
            if special_note:
                notes = special_note
        
        # Explicitly check for "Finish Duty" in the full row - this is important to catch
        if 'finish' in full_row_text and ('duty' in full_row_text or 'dty' in full_row_text):
            duty_mentioned = current_duty_number in full_row_text or f"duty {current_duty_number}" in full_row_text
            # If duty number isn't mentioned but this is the only duty visible on this side, assume it's for this duty
            if duty_mentioned or not re.search(r'duty\s+\d{3}', full_row_text):
                notes = f"Duty {current_duty_number} Finished Duty"
        
        # Check for "Takes Bus" note - common in the running boards and important to preserve
        takes_bus_match = re.search(r'takes\s+bus', full_row_text)
        if takes_bus_match and not notes:
            bus_info = extract_bus_info(full_row_text)
            if bus_info:
                notes = bus_info
        
        # Handle entries based on what we found
        if location or route or arrival or departure or departs_garage or notes:
            # Create the entry
            entry = {
                'location': location,
                'route': route,
                'arrival': arrival,
                'departure': departure,
                'departs_time': departs_time if departs_garage else "",
                'from_loc': next_from_location if next_from_location and location else "",
                'to_loc': location if next_from_location and location else "",
                'notes': notes
            }
            
            # Add non-empty entries to our list
            if any([v for k, v in entry.items() if k not in ['notes']]) or entry['notes']:
                duty_entries.append(entry)
                
                # Update for next row
                if location:
                    next_from_location = location
        
        # If we're on the right side and encounter a different duty, stop processing
        if side == 'right' and row_idx > start_row:
            # Check if this row mentions a different duty
            other_duty_match = re.search(r'duty\s+(\d{3})', side_text)
            if other_duty_match and other_duty_match.group(1) != current_duty_number:
                break
                    
    return duty_entries

def extract_bus_info(text):
    """Extract information about a bus being taken"""
    bus_match = re.search(r'takes\s+bus\s+(\d+)', text)
    if not bus_match:
        return None
    
    bus_number = bus_match.group(1)
    route_match = re.search(r'route\s+(\w+)', text)
    route = route_match.group(1) if route_match else ""
    
    time_match = re.search(r'at\s+(\d{1,2}:\d{2})', text)
    time = time_match.group(1) if time_match else ""
    
    location_match = re.search(r'at\s+\d{1,2}:\d{2}\s+(\w+)', text)
    location = location_match.group(1) if location_match else ""
    
    result = f"Takes Bus {bus_number}"
    if route:
        result += f" (Route {route})"
    if time:
        result += f" at {time}"
    if location:
        result += f" {location}"
    
    return result

def convert_to_csv_format(duties):
    """
    Convert the extracted duty data to the CSV format matching Zone3.
    """
    csv_data = []
    
    for duty in duties:
        duty_num = duty['duty_number']
        reports = duty['reports']
        
        for i, entry in enumerate(duty['entries']):
            # Skip entries with unwanted text in any field
            should_skip = False
            unwanted_texts = [
                "Phibsboro Zone 4 - Routes 9, 122",
                "RUNNING BOARD",
                "BUS ATHA CLIATH"
            ]
            
            for field in ['location', 'route', 'notes', 'from_loc', 'to_loc']:
                if entry[field] and any(text in entry[field] for text in unwanted_texts):
                    should_skip = True
                    break
            
            if should_skip:
                continue
                
            # Determine From/To fields if not already set
            from_loc = entry['from_loc']
            to_loc = entry['to_loc']
            
            if not from_loc and not to_loc and i > 0 and duty['entries'][i-1]['location'] and entry['location']:
                from_loc = duty['entries'][i-1]['location']
                to_loc = entry['location']
            
            csv_data.append({
                'Duty': duty_num,
                'Reports': reports if i == 0 else "",
                'Departs': entry['departs_time'],
                'Location': entry['location'],
                'Route': entry['route'],
                'From': from_loc,
                'To': to_loc,
                'Arrival': entry['arrival'],
                'Departure': entry['departure'],
                'Notes': entry['notes']
            })
    
    return csv_data

if __name__ == "__main__":
    excel_file = "assets/Zone4Boards.xlsx"
    convert_zone4_to_csv(excel_file) 