import pandas as pd
import os
import re
from openpyxl import load_workbook

def convert_zone4_to_csv(excel_file, output_dir=None):
    """
    Convert Zone4 Excel sheets to CSV files matching the Zone3 format.
    Uses a direct approach to extract the duty information.
    """
    # Get base filename
    base_name = os.path.splitext(os.path.basename(excel_file))[0]
    
    # Set output directory
    if output_dir is None:
        output_dir = os.path.dirname(excel_file)
    
    # Ensure output directory exists
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    print(f"Reading Excel file: {excel_file}")
    
    # Load workbook using openpyxl
    wb = load_workbook(filename=excel_file, data_only=True)
    
    for sheet in wb.sheetnames:
        # Skip non-schedule sheets
        if "roster" in sheet.lower() or "summary" in sheet.lower():
            print(f"Skipping sheet '{sheet}' - not a schedule sheet")
            continue
        
        # Determine CSV suffix based on sheet name
        if "M-F" in sheet.upper() or "M-f" in sheet.upper() or "MONDAY" in sheet.upper():
            csv_suffix = "MF"
        elif "SAT" in sheet.upper() or "SATURDAY" in sheet.upper():
            csv_suffix = "Sat"
        elif "SUN" in sheet.upper() or "SUNDAY" in sheet.upper():
            csv_suffix = "Sun"
        else:
            csv_suffix = sheet.replace(" ", "_")
        
        print(f"Processing sheet '{sheet}'...")
        
        # Get the worksheet
        ws = wb[sheet]
        
        # Extract data from the worksheet
        duties = []
        
        # Track the current duty
        current_duty = None
        reports_time = None
        departs_time = None
        duty_entries = []
        
        # Go through each row in the worksheet
        for row in ws.iter_rows(values_only=True):
            # Convert row to strings and make it a flat text for easier searching
            row_values = [str(cell).strip() if cell is not None else "" for cell in row]
            row_text = " ".join(row_values).lower()
            
            # Look for duty number
            duty_match = re.search(r'duty\s+(\d{3})', row_text)
            if duty_match:
                # If we were tracking a duty, save its entries before starting a new one
                if current_duty and duty_entries:
                    duties.append({
                        'duty': current_duty,
                        'reports': reports_time,
                        'entries': duty_entries
                    })
                
                # Start tracking a new duty
                current_duty = duty_match.group(1)
                duty_entries = []
                
                # Look for reports time
                reports_match = re.search(r'reports at\s+(\d{2}:\d{2})', row_text)
                reports_time = reports_match.group(1) if reports_match else ""
                
                continue
            
            # Skip if we're not tracking a duty
            if not current_duty:
                continue
            
            # Skip empty rows
            if not any(row_values):
                continue
            
            # Skip header rows
            if "route" in row_text and "place" in row_text:
                continue
            
            # Look for location, route, times
            location = None
            route = None
            arrival = None
            departure = None
            notes = ""
            
            # Check for location
            for cell in row_values:
                cell_lower = cell.lower()
                if any(loc in cell_lower for loc in ["garage", "charlestown", "limekiln", "psqe", "psqw"]):
                    location = cell
                    break
            
            # Check for route
            for cell in row_values:
                if cell in ["SPL", "9", "122", "Ghost"]:
                    route = cell
                    break
            
            # Check for departs garage
            if "departs garage" in row_text:
                for cell in row_values:
                    if re.match(r'^\d{1,2}:\d{2}(:\d{2})?$', cell):
                        departs_time = cell
                        break
            
            # Look for times
            for cell in row_values:
                if re.match(r'^\d{1,2}:\d{2}(:\d{2})?$', cell):
                    # If we already found a time, assume the second one is departure
                    if arrival:
                        departure = cell
                    else:
                        # For the first time found, assume it's arrival
                        arrival = cell
            
            # Check for "Finish Duty" text
            if "finish" in row_text and "duty" in row_text:
                notes = f"Duty {current_duty} Finished Duty"
            
            # Only add valid entries
            if location or route or arrival or departure:
                entry = {
                    'location': location if location else "",
                    'route': route if route else "",
                    'arrival': arrival if arrival else "",
                    'departure': departure if departure else "",
                    'departs': departs_time if departs_time and location and "garage" in location.lower() else "",
                    'notes': notes
                }
                duty_entries.append(entry)
                
                # Reset departs time after using it
                if departs_time and location and "garage" in location.lower():
                    departs_time = None
        
        # Add the last duty if we were tracking one
        if current_duty and duty_entries:
            duties.append({
                'duty': current_duty,
                'reports': reports_time,
                'entries': duty_entries
            })
        
        # Convert duties to final CSV format
        csv_data = []
        for duty in sorted(duties, key=lambda x: int(x['duty'])):
            duty_num = duty['duty']
            reports = duty['reports']
            
            for i, entry in enumerate(duty['entries']):
                from_loc = ""
                to_loc = ""
                
                # Set From/To if we have consecutive locations
                if i > 0 and duty['entries'][i-1]['location'] and entry['location']:
                    from_loc = duty['entries'][i-1]['location']
                    to_loc = entry['location']
                
                csv_data.append({
                    'Duty': duty_num,
                    'Reports': reports if i == 0 else "",
                    'Departs': entry['departs'],
                    'Location': entry['location'],
                    'Route': entry['route'],
                    'From': from_loc,
                    'To': to_loc,
                    'Arrival': entry['arrival'],
                    'Departure': entry['departure'],
                    'Notes': entry['notes']
                })
        
        # Save to CSV
        if csv_data:
            output_df = pd.DataFrame(csv_data)
            csv_filename = f"{base_name}{csv_suffix}.csv"
            csv_path = os.path.join(output_dir, csv_filename)
            output_df.to_csv(csv_path, index=False)
            print(f"Saved sheet '{sheet}' to {csv_path}")
        else:
            print(f"No duty data extracted from sheet '{sheet}'")
    
    print("Conversion complete!")

if __name__ == "__main__":
    excel_file = "assets/Zone4Boards.xlsx"
    convert_zone4_to_csv(excel_file) 