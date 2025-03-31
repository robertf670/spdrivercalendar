import pandas as pd
import os
import re
from openpyxl import load_workbook
import logging

# Set up logging
logging.basicConfig(level=logging.INFO, 
                    format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def convert_zone4_to_csv(excel_file, output_dir=None):
    """
    Convert Zone4 Excel sheets to CSV files matching the Zone3 format.
    Uses a direct approach with improved accuracy to extract the duty information.
    """
    # Get base filename
    base_name = os.path.splitext(os.path.basename(excel_file))[0]
    
    # Set output directory
    if output_dir is None:
        output_dir = os.path.dirname(excel_file)
    
    # Ensure output directory exists
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    logger.info(f"Reading Excel file: {excel_file}")
    
    # Load workbook using openpyxl
    wb = load_workbook(filename=excel_file, data_only=True)
    
    logger.info(f"Sheets in workbook: {wb.sheetnames}")
    
    for sheet in wb.sheetnames:
        # Skip non-schedule sheets
        if "roster" in sheet.lower() or "summary" in sheet.lower():
            logger.info(f"Skipping sheet '{sheet}' - not a schedule sheet")
            continue
        
        # Determine CSV suffix based on sheet name
        if "M-F" in sheet.upper() or "M-f" in sheet.upper() or "MONDAY" in sheet.upper():
            csv_suffix = "MF"
        elif "SAT" in sheet.upper() or "SATURDAY" in sheet.upper():
            csv_suffix = "Sat" 
        elif "SUN" in sheet.upper() or "SUNDAY" in sheet.upper():
            csv_suffix = "Sun"
        else:
            csv_suffix = sheet.replace(" ", "_")
        
        logger.info(f"Processing sheet '{sheet}' with suffix {csv_suffix}...")
        
        # Process the sheet to extract duty data
        duties = process_sheet(wb[sheet])
        
        if duties:
            logger.info(f"Found {len(duties)} duties in sheet '{sheet}'")
            
            # Convert to CSV format
            csv_data = convert_to_csv_format(duties)
            
            # Create and save the CSV file
            csv_filename = f"{base_name}{csv_suffix}.csv"
            csv_path = os.path.join(output_dir, csv_filename)
            
            # Create DataFrame and save to CSV
            df = pd.DataFrame(csv_data)
            df.to_csv(csv_path, index=False)
            logger.info(f"Saved sheet '{sheet}' to {csv_path}")
        else:
            logger.warning(f"No duty data extracted from sheet '{sheet}'")
    
    logger.info("Conversion complete!")

def process_sheet(worksheet):
    """
    Process a worksheet to extract duty information with improved accuracy.
    This implementation better handles the structure of Zone4 Excel files.
    """
    logger.info(f"Processing worksheet: {worksheet.title}")
    duties = []
    current_duty = None
    current_entries = []
    reports_time = None
    
    # Create a list to store all cells in their original coordinates for better interpretation
    all_cells = []
    for row_idx, row in enumerate(worksheet.iter_rows(values_only=True)):
        row_cells = []
        for col_idx, cell_value in enumerate(row):
            if cell_value is not None:
                row_cells.append({
                    'row': row_idx,
                    'col': col_idx,
                    'value': str(cell_value).strip()
                })
        all_cells.append(row_cells)
    
    logger.info(f"Loaded {len(all_cells)} rows from worksheet")
    
    # Process the content in a more coherent way to identify duties and their structure
    duty_sections = []
    current_section = None
    
    for row_idx, row_cells in enumerate(all_cells):
        # Skip empty rows
        if not row_cells:
            continue
            
        row_text = ' '.join([cell['value'] for cell in row_cells]).lower()
        
        # Check for a duty number
        duty_match = re.search(r'duty\s+(\d{3})', row_text)
        if duty_match:
            # If we found a new duty, save the previous section if any
            if current_section:
                duty_sections.append(current_section)
            
            # Start a new section
            duty_number = duty_match.group(1)
            reports_match = re.search(r'reports at\s+(\d{2}:\d{2})', row_text)
            reports = reports_match.group(1) if reports_match else ''
            
            logger.debug(f"Found duty {duty_number} at row {row_idx}")
            
            current_section = {
                'duty_number': duty_number,
                'reports': reports,
                'start_row': row_idx,
                'rows': [row_cells]
            }
        elif current_section:
            # Add this row to the current section
            current_section['rows'].append(row_cells)
    
    # Add the last section if we have one
    if current_section:
        duty_sections.append(current_section)
    
    logger.info(f"Found {len(duty_sections)} duty sections")
    
    # Process each duty section to extract structured information
    for section in duty_sections:
        duty_number = section['duty_number']
        reports = section['reports']
        entries = []
        
        # Find all locations in this duty
        locations = []
        for row_cells in section['rows']:
            for cell in row_cells:
                cell_value = cell['value'].lower()
                if any(loc in cell_value for loc in ["garage", "charlestown", "limekiln", "psqe", "psqw"]):
                    locations.append({
                        'row': cell['row'],
                        'col': cell['col'],
                        'value': cell['value']
                    })
        
        # Find all times in this duty
        times = []
        for row_cells in section['rows']:
            for cell in row_cells:
                if re.match(r'^\d{1,2}:\d{2}(:\d{2})?$', cell['value']):
                    times.append({
                        'row': cell['row'],
                        'col': cell['col'],
                        'value': cell['value']
                    })
        
        # Find all routes in this duty
        routes = []
        for row_cells in section['rows']:
            for cell in row_cells:
                if cell['value'] in ["SPL", "9", "122", "Ghost"]:
                    routes.append({
                        'row': cell['row'],
                        'col': cell['col'],
                        'value': cell['value']
                    })
        
        # Look for "Departs Garage" rows
        departs_garage_rows = []
        for i, row_cells in enumerate(section['rows']):
            row_text = ' '.join([cell['value'] for cell in row_cells]).lower()
            if "departs garage" in row_text:
                departs_time = None
                for cell in row_cells:
                    if re.match(r'^\d{1,2}:\d{2}(:\d{2})?$', cell['value']):
                        departs_time = cell['value']
                        break
                
                if departs_time:
                    departs_garage_rows.append({
                        'row_idx': section['start_row'] + i,
                        'departs_time': departs_time
                    })
        
        logger.debug(f"Duty {duty_number}: Found {len(locations)} locations, {len(times)} times")
        
        # Map locations to their rows for quicker lookups
        location_rows = {}
        for loc in locations:
            location_rows[loc['row']] = loc['value']
        
        # Process each row in this duty section
        for i, row_cells in enumerate(section['rows']):
            row_idx = section['start_row'] + i
            row_text = ' '.join([cell['value'] for cell in row_cells]).lower()
            
            # Skip rows with no content
            if not row_cells:
                continue
            
            # Get location for this row
            location = location_rows.get(row_idx, "")
            
            # Extract times on this row
            row_times = [t for t in times if t['row'] == row_idx]
            row_times.sort(key=lambda x: x['col'])  # Sort by column to maintain order
            
            arrival = None
            departure = None
            
            # Handle times based on row context
            if "departs garage" in row_text and row_times:
                # For "Departs Garage" rows, the time is the departs time
                departs_time = row_times[0]['value'] if row_times else None
                
                # If there's a second time, it's likely a scheduled finish time
                if len(row_times) > 1:
                    departure = row_times[1]['value']
            elif row_times:
                # For other rows, first time is arrival, second is departure
                arrival = row_times[0]['value'] if row_times else None
                if len(row_times) > 1:
                    departure = row_times[1]['value']
            
            # Get route for this row
            route = None
            for r in routes:
                if r['row'] == row_idx:
                    route = r['value']
                    break
            
            # Check for "Departs Garage" time
            departs_time = None
            for dg in departs_garage_rows:
                if dg['row_idx'] == row_idx:
                    departs_time = dg['departs_time']
                    break
            
            # Check for notes
            notes = ""
            if "finish" in row_text and "duty" in row_text and f"duty {duty_number}" in row_text:
                notes = f"Duty {duty_number} Finished Duty"
            
            # Only add entries with useful information
            if location or route or arrival or departure:
                entry = {
                    'location': location,
                    'route': route if route else "",
                    'arrival': arrival if arrival else "",
                    'departure': departure if departure else "",
                    'departs_time': departs_time if departs_time else "",
                    'notes': notes
                }
                entries.append(entry)
                logger.debug(f"Added entry for duty {duty_number}: {entry}")
        
        # Only add duties with entries
        if entries:
            duties.append({
                'duty_number': duty_number,
                'reports': reports,
                'entries': entries
            })
            logger.debug(f"Added duty {duty_number} with {len(entries)} entries")
    
    # Sort duties by duty number
    duties.sort(key=lambda x: int(x['duty_number']))
    
    return duties

def convert_to_csv_format(duties):
    """
    Convert the extracted duty data to the CSV format matching Zone3.
    """
    csv_data = []
    
    for duty in duties:
        duty_num = duty['duty_number']
        reports = duty['reports']
        
        for i, entry in enumerate(duty['entries']):
            # Determine From/To fields
            from_loc = ""
            to_loc = ""
            
            if i > 0 and duty['entries'][i-1]['location'] and entry['location']:
                from_loc = duty['entries'][i-1]['location']
                to_loc = entry['location']
            
            csv_data.append({
                'Duty': duty_num,
                'Reports': reports if i == 0 else "",
                'Departs': entry['departs_time'] if entry['location'] and "garage" in entry['location'].lower() else "",
                'Location': entry['location'],
                'Route': entry['route'],
                'From': from_loc,
                'To': to_loc,
                'Arrival': entry['arrival'],
                'Departure': entry['departure'],
                'Notes': entry['notes']
            })
    
    return csv_data

if __name__ == "__main__":
    excel_file = "assets/Zone4Boards.xlsx"
    convert_zone4_to_csv(excel_file) 